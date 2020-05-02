"""
    Author：Tiger Z
    Time  ：2020-4-20 10:58:07
    Log   ：
        V0.1：
            初版，收集openpyxl库对Excel基本操作


"""
# 导入openpyxl package
import openpyxl

# 打开一个已有工作簿
f = openpyxl.load_workbook('test.xlsx')

# sheet相关操作
# 获取sheet名称列表
name_list = f.sheetnames
name_list = f.get_sheet_names()

# 通过名称调用sheet
table = f['name']
table = f.get_sheet_by_name('name')

# 通过检索调用sheet
table = f.get_sheet_by_name(sheet_names[index])

# 调用正在运行的sheet，一般为最后一个
table = f.active

# 改变sheet的名字
table.title = 'newname'

# 单元格相关操作

# 读取单元格或范围切片
c = table['A1']  # 获取'A1'，返回class
c = table.cell(row=1, column=1)  # 按行列数获读取

row5 = table[5]  # 获取第5行，返回元组
colA = table['C']  # 获取C列
cell_range = table['A1':'B4']  # 获取范围切片元组
row_range = table[5:8]  # 5到8行
col_range['A:D']  # A到D列

# 按行列数读取范围切片
for row in table.iter_rows(min_row=1, max_col=3, max_row=2):  # 按行读取
    for col in table.iter_cols(min_row=1, max_col=3, max_row=2):  # 按列读取

table.max_row  # 最大行数
table.max_column  # 最大列数
table.rows  # 按行遍历
table.columns  # 按列遍历

# 读取单元格的值
c = table['A2'].value  # 按位置读取
c = table.cell(row=1, column=1).value  # 按行列数读取

# 需要注意的是openpyxl中按行列数检索时参数名‘row =’及‘column =’ 不能省略，而且均从1开始计数，这与xlrd有所不同。
####一个例子
# 利用openpyxl读出图1所示表中的一些信息，代码及结果如下：

import openpyxl  # 导入openpyxl

f = openpyxl.load_workbook('demo.xlsx')  # 打开工作簿
print(f.sheetnames)  # 打印sheet名称列表
table = f['成绩单']  # 调用成绩单sheet
print(table.max_column)  # 打印最大列数
print(table.cell(row=3, column=3).value)  # 打印C3的值
for i in table['A']:  # 依次打印A列的值
    print(i.value)

import datetime
from openpyxl import Workbook

wb = Workbook()
print(type(wb))
# grab the active worksheet
ws = wb.active
print(type(ws))
# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")

# 创建一个工作表，然后
wb = Workbook()

# 找到活动的sheet页。空的excel表默认的sheet页就叫Sheet，
sheet = wb.active

#可以直接给title属性赋值进行重命名。
# 这个属性是可读可写的。当然，这个只针对当前活动页，别的页的话，可以用create_sheet和remove_sheet进行添加和删除。
sheet.title = "New Shit"

# 往sheet页里面写内容就比较简单了，跟上面读一样，
sheet['C3'] = 'Hello world!'

for i in range(10):
  sheet["A%d" % (i+1)].value = i + 1

# 我们还可以进行花式操作，比如写写公式：
sheet["E1"].value = "=SUM(A:A)"

# 最后记得保存
wb.save('保存一个新的excel.xlsx')

# 打开一个xlsx文件。
wb = load_workbook("template.xlsx")

print(wb.sheetnames)    # ['Sheet1', 'Sheet2', 'Sheet3']
# 可以看看打开的excel表里面有哪些sheet页。
# sheet = wb.get_sheet_by_name("Sheet3")
# 读取到指定的Sheet页，sheet就变得神奇了，想要的内容都在这里。比如：
print(sheet["C"])    # (<Cell Sheet3.C1>, <Cell Sheet3.C2>, <Cell Sheet3.C3>, <Cell Sheet3.C4>, <Cell Sheet3.C5>, <Cell Sheet3.C6>, <Cell Sheet3.C7>, <Cell Sheet3.C8>, <Cell Sheet3.C9>, <Cell Sheet3.C10>)      <-第C列
print(sheet["4"])    # (<Cell Sheet3.A4>, <Cell Sheet3.B4>, <Cell Sheet3.C4>, <Cell Sheet3.D4>, <Cell Sheet3.E4>)     <-第4行
print(sheet["C4"].value)    # c4     <-第C4格的值
print(sheet.max_row)    # 10     <-最大行数
print(sheet.max_column)    # 5     <-最大列数
for i in sheet["C"]:
  print(i.value, end=" ")    # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值

  # 获取单元格
  # 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = sheet['B4']
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}')  # 返回的数字就是int型

# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = sheet.cell(row=4, column=2)
print(b4_too.value)