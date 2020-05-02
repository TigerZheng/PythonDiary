"""
    Author：Tiger Z
    Time  ：2020-4-20 10:58:07
    Log   ：
        v0.1：
            根据从执行文件读入的参数，解二元一次方程组

"""
import numpy as np
from openpyxl import load_workbook
"""
Demo:
x + 2y = 3
4x ＋ 5y = 6
        A = np.mat('1,2; 4,5')    # 构造系数矩阵 A
        b = np.mat('3,6').T       # 构造转置矩阵 b （这里必须为列向量）
        r = np.linalg.solve(A,b)  # 调用 solve 函数求解
        print r
Out[1]: [[-1.]
         [ 2.]]
         
        x = -1
        y = 2 
"""
def FuelRes2LiterFactor():
    """"
        功能：
            计算燃油电阻转换为升数的系数
    """
    A = np.mat("1,308;1,279")
    b = np.mat("8000,12000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,279;1,249")
    b = np.mat("12000,17000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,249;1,226")
    b = np.mat("17000,22000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,226;1,197")
    b = np.mat("22000,28000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,197;1,168")
    b = np.mat("28000,33000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,168;1,138")
    b = np.mat("33000,38000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,138;1,109")
    b = np.mat("38000,43000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,109;1,79")
    b = np.mat("43000,48000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    # list_A = []
    # list_b = []
    # cnt = 0
    # wb = load_workbook("test.xlsx")
    # asheet = wb["Sheet1"]
    # # 获取 方程组系数
    # for i in asheet['A']:
    #     if i.value != None:
    #         list_A.append(i.value)
    #         cnt = cnt+1
    # # 获取 方程组右侧的值
    # for i in asheet['B']:
    #     if i.value != None:
    #         list_b.append(i.value)
    #
    # print(len(list_A))
    # for i in range((len(list_A)-1)):
    #     # A = np.mat("1,2;4,5")
    #     A = np.mat("list_A[i],1;list_A[i+1],1")
    #     # b = np.mat("3,6").T
    #     b = np.mat("listb[i];list_b[i+1]").T
    #     r= np.linalg.solve(A,b)
    #     np.savetxt('factor.txt', r, delimiter=',')
    # # file = open("factor.txt", "a")
    # # file.write(r + "\n")
    # # file.close()

def FuelRes2LiterFactor_190531():
    """"
        功能：
            计算燃油电阻转换为升数的系数
    """
    A = np.mat("1,322;1,294")
    b = np.mat("5000,10000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,294;1,271")
    b = np.mat("10000,14000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,271;1,241")
    b = np.mat("14000,19000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,241;1,200")
    b = np.mat("19000,24000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,200;1,168")
    b = np.mat("24000,33000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,168;1,138")
    b = np.mat("33000,38000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,138;1,94")
    b = np.mat("38000,45000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,94;1,79")
    b = np.mat("45000,48000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')


def FuelRes2LiterFactor_190612():
    """"
        功能：
            计算燃油电阻转换为升数的系数
    """
    A = np.mat("1,322;1,286")
    b = np.mat("5000,11000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,286;1,249")
    b = np.mat("11000,17000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,249;1,226")
    b = np.mat("17000,22000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,226;1,197")
    b = np.mat("22000,28000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,197;1,168")
    b = np.mat("28000,33000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,168;1,138")
    b = np.mat("33000,38000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,138;1,94")
    b = np.mat("38000,45000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')

    A = np.mat("1,94;1,79")
    b = np.mat("45000,48000").T
    r= np.linalg.solve(A,b)
    with open('factor.txt', 'ab') as f:
        np.savetxt(f, r, delimiter=',',fmt='%.5f')


def FuelResCal():
    """"
        功能：
            计算燃油电阻阻值
            F70项目
    """
    list_IG = []
    list_Fuel = []

    wb = load_workbook("fuel.xlsx")
    asheet = wb["Sheet1"]

    # 获取 IG电压数据
    for i in asheet['A']:
        if i.value != None:
            list_IG.append(i.value)

    # 获取 燃油电阻测试端电压数据
    for i in asheet['B']:
        if i.value != None:
            list_Fuel.append(i.value)

    for i in range(3):  # 此处需要根据实际情况进行修改，后期优化
        res = 8000000*list_Fuel[i]/(2457*list_IG[i]-6680*list_Fuel[i])+30
        file = open("res.txt", "a")
        res_str = str(res)
        # print(type(res_str))
        file.write(res_str + "\n")
        file.close()



def main():
    # FuelRes2LiterFactor()         # 计算阻值转换为升数的参数
    # FuelResCal()                    # 计算燃油电阻的值
    #FuelRes2LiterFactor_190531()    # 20190531 新段系数计算
    FuelRes2LiterFactor_190612()

if __name__ == "__main__":
    main()
