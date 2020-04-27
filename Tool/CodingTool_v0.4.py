"""
    Author：Tiger Z
    Time  ：2019-4-2 15:49:41
    Log   ：
        V0.1：
            初版，初步实现LED灯模块的代码生成，代码内容还需优化以满足目标格式
        
        V0.2：
            添加CAN信号相关关代码生成工具
        
        V0.3:
            2020-4-15 15:58:19
            添加AID及LID代码生成工具
            
        V0.4
            2020-4-27 15:17:47
            添加 MenuFSM框架代码生成工具
"""
from typing import List, Any

from openpyxl import load_workbook


def CodeWriter_LID():
    """"
        LID 模块代码自动生成器
        2020-4-14 16:32:03
        基本功能实现：从电子表格读取数据，按照指定格式写入txt中
    """
    # 需求ID
    list_id = [] 
    # 功能名称   
    list_name = []
    # 引脚
    list_pin = []
    # 上拉条件
    list_pullup = []
    # 采集结果记录变量
    list_value = []
    # 有效触发电平
    list_triger = []

    # 打开指定的Excel
    wb = load_workbook("LID.xlsx")

    # 通过名称调用指定sheet
    asheet = wb["Sheet1"]

    # 获取 ID
    for i in asheet['A']:
        if i.value is not None:
            list_id.append(i.value)

    # 获取 name
    for i in asheet['B']:
        if i.value is not None:
            list_name.append(i.value)

    # 获取 pin
    for i in asheet["C"]:
        if i.value is not None:
            list_pin.append(i.value)

    # 获取 pullup
    for i in asheet["D"]:
        if i.value is not None:
            list_pullup.append(i.value)

    # 获取 value
    for i in asheet["E"]:
        if i.value is not None:
            list_value.append(i.value)

    # 获取 triger
    for i in asheet["F"]:
        if i.value is not None:
            list_triger.append(i.value)

    file1 = open("LIDCode.txt", "a")
    for i in list_id:
        file1.write("/******************************************************************************************/\n")
        file1.write("//    ID:  " + str(list_id[i]) + "    \n")
        file1.write("/******************************************************************************************/\n")
        file1.write("void LID_" + list_name[i] + "_ValidFun(void *pData)\n")
        file1.write("{\n")
        file1.write("    " + list_value[i] + " = TRUE;\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void LID_" + list_name[i] + "_InValidFun(void *pData)\n")
        file1.write("{\n")
        file1.write("    " + list_value[i] + " = FALSE;\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("LIDetect_State_Obj LID_" + list_name[i] + "_State;\n")
        file1.write("const LIDetect_Set_Obj LID_" + list_name[i] + "_Set = ;\n")
        file1.write("{\n")
        file1.write("    " + list_pin[i] + ",\n")
        file1.write("    " + list_pullup[i] + ",\n")
        file1.write("    LID_SET_PINVALID_KEY,\n")
        file1.write("    NULL,\n")
        file1.write("    20,//ms\n")
        file1.write("    " + list_triger[i] + ",\n")
        file1.write("    0,\n")
        file1.write("    LID_" + list_name[i] + "_ValidFun,\n")
        file1.write("    0,\n")
        file1.write("    LID_" + list_name[i] + "_InValidFun,\n")
        file1.write("    0,\n")
        file1.write("    &LID_" + list_name[i] + "_State\n")
        file1.write("}\n")
    file1.close()


def CodeWriter_AID():
    """"
        AID 模块代码自动生成器
        2020-4-15 14:48:45
        基本功能实现：从电子表格读取数据，按照指定格式写入txt中
    """
    # 需求ID
    list_id = []
    # 需求名称
    list_name = []
    # 引脚定义
    list_pin = []
    # 采样周期
    list_SampleCycle = []
    # 采样数据buffer大小
    list_bufsize = []
    # 存储信息的变量
    list_value = []

    # 打开指定的Excel
    wb = load_workbook("AID.xlsx")

    # 通过名称调用指定sheet
    asheet = wb["Sheet1"]

    # 获取 ID
    for i in asheet['A']:
        if i.value != None:
            list_id.append(i.value)

    # 获取 name
    for i in asheet['B']:
        if i.value != None:
            list_name.append(i.value)

    # 获取 pin
    for i in asheet["C"]:
        if i.value != None:
            list_pin.append(i.value)

    # 获取 SampleCycle
    for i in asheet["D"]:
        if i.value != None:
            list_SampleCycle.append(i.value)

    # 获取 bufsize
    for i in asheet["E"]:
        if i.value != None:
            list_bufsize.append(i.value)

    # 获取 value
    for i in asheet["F"]:
        if i.value != None:
            list_value.append(i.value)

    file1 = open("AID.txt", "a")
    for i in list_id:
        file1.write("/******************************************************************************************/\n")
        file1.write("//    ID:  " + str(list_id[i]) + "    \n")
        file1.write("/******************************************************************************************/\n")
        file1.write("void AID_" + list_name[i] + "_Fun(void *pdata, u16 res)\n")
        file1.write("{\n")
        file1.write("    " + list_value[i] + " =0;\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("AIDetect_State_Obj AID_" + list_name[i] + "_State;\n")
        file1.write("#define AID_" + list_name[i] + "_BUF_NUM ("+ str(list_bufsize[i]) + "u)\n")
        file1.write("u16 AID_" + list_name[i] + "_BUF["+ str(list_bufsize[i]) + "+1] = {0};\n")
        file1.write("\n")
        file1.write("AIDetect_Set_Obj AID_" + list_name[i] + "_Set = \n")
        file1.write("{\n")
        file1.write("    " + list_pin[i] + ",\n")
        file1.write("    AID_" + list_name[i] + "_BUF,\n")
        file1.write("    " + str(list_SampleCycle[i]) + ",\n")
        file1.write('    AID_SET_DEALMODE_MAXMIN_AVERAGE,\n')
        file1.write("    AID_" + list_name[i] + "_BUF_NUM,\n")
        file1.write("    1,\n")
        file1.write("    AID_" + list_name[i] + "_BUF_NUM-1,\n")
        file1.write("    AID_" + list_name[i] + "_Fun,\n")
        file1.write("    0,\n")
        file1.write("    &AID_" + list_name[i] + "_State\n")
        file1.write("}\n")
        file1.write("\n")
    file1.close()


def CodeWriter_LED():
    """"
        LED 模块代码自动生成器
        2019-4-2 15:52:29
        基本功能实现：从电子表格读取数据，按照指定格式写入txt中
    """
    list_ID = []
    list_name = []
    list_condition = []

    # 打开指定的Excel
    wb = load_workbook("LED.xlsx")

    # 通过名称调用指定sheet
    asheet = wb["Sheet1"]

    # 获取 ID
    for i in asheet['A']:
        if i.value is not None:
            list_ID.append(i.value)

    # 获取 name
    for i in asheet['C']:
        if i.value is not None:
            list_name.append(i.value)

    # 获取 condition
    for i in asheet["D"]:
        if i.value is not None:
            list_condition.append(i.value)

    file1 = open("Ledcode.txt", "a")

    for i in list_ID:
        file1.write("/******************************************************************************************/\n")
        file1.write("//    ID:  " + str(list_ID[i]) + "    \n")
        file1.write("/******************************************************************************************/\n")
        file1.write("u8 LED_GetState_" + list_name[i] + "(void *pData)\n")
        file1.write("{\n")
        file1.write(" u8  " + list_condition[i] + " =1;\n")
        file1.write("   if( " + list_condition[i] + " == 1)\n")
        file1.write("   {\n")
        file1.write("       return LED_LightStatus_On;\n")
        file1.write("   }\n")
        file1.write("   else\n")
        file1.write("   {\n")
        file1.write("       return LED_LightStatus_Off;\n")
        file1.write("   }\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void LED_On_" + list_name[i] + "(void)\n")
        file1.write("{\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void LED_Off_" + list_name[i] + "(void)\n")
        file1.write("{\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("\n")
    file1.close()


def CodeWriter_CANMsg():
    """"
        CAN通信部分代码生成函数
        时间：2019-5-22 14:13:50
        基本功能实现：从电子表格读取数据，按照指定格式写入txt中
    """
    CAN_ID = []
    CAN_Msg = []
    CAN_Signal = []
    CAN_Byte = []
    CAN_StartBit = []
    # CAN_EndBit = []
    CAN_BitLen = []
    TotalMsgNum = 0

    # 打开指定的Excel
    wb = load_workbook("baowen.xlsx")

    # 通过名称调用sheet
    asheet = wb["Sheet1"]

    # 获取 ID
    for i in asheet['C']:
        if i.value is not None:
            CAN_ID.append(i.value)

    # 获取信号帧名称
    for i in asheet['A']:
        if i.value is not None:
            CAN_Msg.append(i.value)
            TotalMsgNum = TotalMsgNum + 1

    # 获取 CAN信号
    for i in asheet["G"]:
        if i.value is not None:
            CAN_Signal.append(i.value)

    # 获取 CAN信号所在字节
    for i in asheet["I"]:
        if i.value is not None:
            CAN_Byte.append(i.value)

    # 获取 CAN信号起始bit
    for i in asheet["J"]:
        if i.value is not None:
            CAN_StartBit.append(i.value)
            CAN_Byte.append((i.value)/8)

    # 获取 CAN信号长度
    for i in asheet["J"]:
        if i.value is not None:
            CAN_BitLen.append(i.value)

    # 获取 CAN信号结束bit
    # for i in asheet["C"]:
    #     if i.value is not None:
    #         CAN_EndBit.append(i.value)

    # 生成 CanMsgStruct结构体
    file1 = open("CANMsgStruct.txt", "a")

    file1.write("typedef struct  __Can_Variable_Struct \n")
    file1.write("{ \n")

    for i in range(TotalMsgNum):

        if(int(CAN_BitLen[i]) <= 8):
            file1.write("    u8  " + CAN_Msg[i] + "_" + CAN_Signal[i] + ";\n")
        elif(int(CAN_BitLen[i]) <= 16):
            file1.write("    u16  " + CAN_Msg[i] + "_" + CAN_Signal[i] + ";\n")
        elif(int(CAN_BitLen[i]) <= 32):
            file1.write("    u32  " + CAN_Msg[i] + "_" + CAN_Signal[i] + ";\n")

    file1.write("}Can_Variable_Struct;\n")
    file1.close()


def ExcelMsg_Filter():
    """"
        Excel 指定关键字过滤
        2019-9-23 16:06:26
        基本功能实现：按照指定关键字过滤电子表格中的信息
    """
    list_ID = []
    # 打开指定的Excel
    wb = load_workbook("1.xlsx")
    # 通过名称调用sheet
    asheet = wb["Sheet1"]
    file1 = open(".txt", "a")
    for i in asheet['E']:
        if i.value is not None:
            if "MCU" in i.value:
                msg = i.value
                file1.write( msg)

    file1.close()


def CodeWriter_MenuFSMFrame():
    """"
        MenuFSM 模块框架代码自动生成器
        2020-4-27 11:23:47
        基本功能实现：从电子表格读取数据，按照指定格式写入txt中
    """
    # 序号
    list_id = []
    # 功能名称
    list_name = []
    # 注释
    list_comment = []

    # 当前菜单条目数
    list_item = []

    # 打开指定的Excel
    wb = load_workbook("MenuFSM.xlsx")

    # 通过名称调用指定sheet
    asheet = wb["Sheet1"]

    # 序号
    for i in asheet['A']:
        if i.value is not None:
            list_id.append(i.value)

    # 获取 name
    for i in asheet['B']:
        if i.value is not None:
            list_name.append(i.value)

    # 获取 注释信息
    for i in asheet["C"]:
        if i.value is not None:
            list_comment.append(i.value)

    # 获取 注释信息
    for i in asheet["D"]:
        if i.value is not None:
            list_item.append(i.value)

    file1 = open("MenuFSMCode.txt", "a")
    # 创建函数
    for i in list_id:
        file1.write("/******************************************************************************************/\n")
        file1.write("//    ID:  " + list_comment[i] + "    \n")
        file1.write("/******************************************************************************************/\n")
        file1.write("void MenuFSM_" + list_name[i] + "_ScreenInit(void)\n")
        file1.write("{\n")
        file1.write("    MenuScreenInit("+str(list_item[i])+");\n")
        file1.write("    Menu_PUSH();\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void MenuFSM_" + list_name[i] + "_SetKey(void)\n")
        file1.write("{\n")
        file1.write("    switch(Item)\n")
        file1.write("    {\n")
        file1.write("        case 0:\n")
        file1.write("            Menu_Jump2Menu(SelfMenuIndex, FlashMode_AutoInit);\n")
        file1.write("            break;\n")
        file1.write("\n")
        file1.write("        default:\n")
        file1.write("            Menu_Jump2Menu(SelfMenuIndex, FlashMode_NoAction);\n")
        file1.write("            break;\n")
        file1.write("     }\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void MenuFSM_" + list_name[i] + "_DownKey(void)\n")
        file1.write("{\n")
        file1.write("    Menu_RightKey();\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void MenuFSM_" + list_name[i] + "_UpKey(void)\n")
        file1.write("{\n")
        file1.write("    Menu_LeftKey();\n")
        file1.write("}\n")
        file1.write("\n")
        file1.write("void MenuFSM_" + list_name[i] + "_BackKey(void)\n")
        file1.write("{\n")
        file1.write("    Menu_POP();\n")
        file1.write("}\n")
        file1.write("\n")

    # 创建KeyTab []
    file1.write("const KbdTabStruct KeyTab[] =\n")
    file1.write("{\n")
    for i in list_id:
        file1.write("    {"+str(list_id[i]*5+0)+", "+str(list_id[i]*5+1)+", "+str(list_id[i]*5+2)+", "+str(list_id[i]*5+3)+", "+str(list_id[i]*5+4)+", "+"MenuFSM_" + list_name[i] + "_ScreenInit},\n")
        file1.write("    {"+str(list_id[i]*5+1)+", 0, 0, 0, 0,"+"MenuFSM_" + list_name[i] + "_SetKey},\n")
        file1.write("    {"+str(list_id[i]*5+2)+", 0, 0, 0, 0,"+"MenuFSM_" + list_name[i] + "_DownKey},\n")
        file1.write("    {"+str(list_id[i]*5+3)+", 0, 0, 0, 0,"+"MenuFSM_" + list_name[i] + "_UpKey},\n")
        file1.write("    {"+str(list_id[i]*5+4)+", 0, 0, 0, 0,"+"MenuFSM_" + list_name[i] + "_BackKey},\n")
        file1.write("\n")
    file1.write("};\n")
    file1.close()


def main():
    # 生成LID模块代码
    # CodeWriter_LID()
    # 生成AID模块代码
    # CodeWriter_AID()
    # 生成LED模块代码
    # CodeWriter_LED()
    # CAN 通信模块代码生成
    # CodeWriter_CANMsg()
    # Excel 操作工具
    # ExcelMsg_Filter()
    # MenuFSM 模块代码生成
    CodeWriter_MenuFSMFrame()


if __name__ == "__main__":
    main()
